import openpyxl
import pandas as pd

from constants import *
from openpyxl.styles import Font


class DataResultWriter:
    def __init__(self, result_df: pd.DataFrame):
        self.__green_fill = openpyxl.styles.PatternFill(start_color=GREEN, end_color=GREEN,
                                                        fill_type='solid')
        self.__orange_headers_fill = openpyxl.styles.PatternFill(start_color=ORANGE_HEADERS, end_color=ORANGE_HEADERS,
                                                                 fill_type='solid')
        self.__orange_fill = openpyxl.styles.PatternFill(start_color=ORANGE, end_color=ORANGE,
                                                         fill_type='solid')
        self.__sky_blue_fill = openpyxl.styles.PatternFill(start_color=SKY_BLUE, end_color=SKY_BLUE,
                                                           fill_type='solid')
        self.__blue_fill = openpyxl.styles.PatternFill(start_color=BLUE, end_color=BLUE,
                                                       fill_type='solid')
        self.__bold_font = Font(bold=True)
        self.__result_df = result_df
        self.__result_df.to_excel(RESULT_FILENAME, engine='openpyxl', index=False)
        self.__workbook: openpyxl.workbook = openpyxl.load_workbook(RESULT_FILENAME)
        self.__sheet = self.__workbook.active

    def write_result(self):
        self.__set_sheet_title()
        self.__set_filtered_rows()
        self.__set_alignment_cells()
        self.__set_cells_width()
        self.__set_first_row_color()
        self.__set_all_borders()
        self.__set_cells_color()
        self.__workbook.save(RESULT_FILENAME)

    def __set_cells_color(self):
        columns_fill = {
            'A': self.__sky_blue_fill,
            'B': self.__sky_blue_fill,
            'C': self.__orange_fill,
            'D': self.__green_fill,
            'E': self.__blue_fill,
            'F': self.__blue_fill
        }

        for column_letter, fill_color in columns_fill.items():
            for cell in self.__sheet[column_letter][1:]:
                cell.fill = fill_color

                # Для столбцов 'E' и 'F' также проверяем наличие подстроки "ОК"
                if column_letter in ['E', 'F']:
                    if "ОК" in cell.value:
                        cell.fill = self.__blue_fill
                    else:
                        cell.font = self.__bold_font
                        cell.fill = self.__orange_fill

                # Для столбца 'C' проверяем соответствие значений с 'B'
                if column_letter == 'C':
                    input_value = cell.offset(column=-1, row=0).value
                    if input_value != cell.value:
                        cell.font = self.__bold_font
                        cell.fill = self.__orange_fill
                    else:
                        cell.fill = self.__green_fill

    def __set_all_borders(self):
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )

        for row in self.__sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

    def __set_first_row_color(self):
        for cell in self.__sheet[1]:
            cell.fill = self.__orange_headers_fill

    def __set_cells_width(self):
        # Получение максимальной длины содержимого для каждого столбца
        max_lengths = {}
        for row in self.__sheet.iter_rows():
            for cell in row:
                if cell.value:
                    length = len(str(cell.value))
                    max_lengths[cell.column_letter] = max(max_lengths.get(cell.column_letter, 0), length)

        # Установка ширины каждого столбца на основе максимальной длины его содержимого
        for col, length in max_lengths.items():
            width = int(1.4 * length)
            self.__sheet.column_dimensions[col].width = width

    def __set_filtered_rows(self):
        self.__sheet.auto_filter.ref = self.__sheet.dimensions

    def __set_sheet_title(self):
        self.__sheet.title = RESULT_SHEET_NAME

    def __set_alignment_cells(self):
        # Установка выравнивания для первой строки
        for cell in self.__sheet[1]:
            cell.alignment = openpyxl.styles.Alignment(horizontal='left')

        # Установка выравнивания для всех остальных строк
        for row in self.__sheet.iter_rows(min_row=2, max_row=self.__sheet.max_row):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
